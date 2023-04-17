from tkinter import*
from tkinter import ttk
import tkinter.messagebox as mb
import datetime

import openpyxl
import pymysql
import tkinter as tk

def addz():
    a = txt_one.get()
    b = txt_two.get()
    c = txt_fr.get()
    d = txt_for.get()
    v = txt_fv.get()
    g = txt_sx.get()
    st = "\n" + a + "\n" + b + "\n" + c + "\n" + d + "\n" + v + "\n" + g + "\n"

    with open('zhurnal.txt', 'a', encoding='utf-8') as file:
        file.write(st)
    mb.showinfo("ok", "sucsessfull")

def add_date():

    data = txt_one.get()
    st = data + "\n"
    with open('date.txt', 'a', encoding='utf-8') as file:
        file.write(st)
    mb.showinfo("ok", "sucsessfull")
    wb = openpyxl.load_workbook(filename='Otchet.xlsx')

def excel():
    data = txt_one.get()
    wb = openpyxl.load_workbook('Otchet.xlsx')
    sheet = wb.active
    print('Данные импортированы успешно!')
    for i in range(1, 2):
        a1 = sheet[f'A1'].value
        b1 =sheet[f'B1'].value
        c1 = sheet[f'C1'].value
        d1 = sheet[f'D1'].value
        e1 = sheet[f'E1'].value
        f1 = sheet[f'F1'].value
        g1 = sheet[f'G1'].value
        h1 = sheet[f'H1'].value
        i1 = sheet[f'I1'].value
        j1 = sheet[f'J1'].value
        k1 = sheet[f'K1'].value
        print(a1)
        if a1 == None:
            c99 = sheet.cell(row=1, column=1)
            c99.value = data
        elif b1 == None:
            c99 = sheet.cell(row=1, column=2)
            c99.value = data
    # Максимум 11 значений


    mb.showinfo('ok', 'успешно')
    wb.save(f"Otchet.xlsx")




#Авторизация
root1 = Tk()
root1.title("Авториация")
root1.geometry('290x400')
root1.configure(background='#66CDAA')
#Лейблы
lbl_one0 = Label(root1, text='Дата',bg='#66CDAA').grid(row=6, column=5)
lbl_one1 = Label(root1, text='Номер зачетки',bg='#66CDAA').grid(row=8, column=5)
lbl_one1 = Label(root1, text='ФИО',bg='#66CDAA').grid(row=10, column=5)
lbl_one1 = Label(root1, text='Группа',bg='#66CDAA').grid(row=12, column=5)
lbl_one1 = Label(root1, text='Дисциплина',bg='#66CDAA').grid(row=14, column=5)
lbl_one1 = Label(root1, text='Время сдачи',bg='#66CDAA').grid(row=16, column=5)
#Поля для ввода
txt_one = Entry(root1, width=20)
txt_one.grid(row=7, column=5)
txt_two = Entry(root1, width=20)
txt_two.grid(row=9, column=5)
txt_fr = Entry(root1, width=20)
txt_fr.grid(row=11, column=5)
txt_for = Entry(root1, width=20)
txt_for.grid(row=13, column=5)
txt_fv = Entry(root1, width=20)
txt_fv.grid(row=15, column=5)
txt_sx = Entry(root1, width=20)
txt_sx.grid(row=17, column=5)
#Поля для ввода
#Кнопки
btn_save = Button(text='Отпечатать даты', command=add_date).grid(row=25, column=5, padx= 100, pady = 5)
btn_reg = Button(text='Показать в Excel', command=excel).grid(row=26, column=5, pady = 5)
btn_mng = Button(text='Добавить запись', command=addz).grid(row=27, column=5, pady = 5)
#Кнопки

root1.mainloop()